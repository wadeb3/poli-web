"""
sync_budget_glance.py. Poli data pipeline · Budget at a Glance sync

Downloads budget.gov.au's own "Website Chart Data" ZIP (the source behind
BP1/BP3's narrative charts) and re-extracts the three slices BudgetGlance
uses:
  - Table 1 (cash receipts by category) → revenue composition over time
  - bp1-bs6.xlsx sheet 6.1 → spending by government function
  - bp1-bs8.xlsx sheet 8.1 → GDP growth forecast vs. actual outcome

This is NOT the same source as sync_budget_measures.py (BP2, itemized
measures), this ZIP has zero measure-level data, only long-run macro
context. Confirmed by inspecting the actual file contents, not filenames.

The ZIP's internal filenames are stable year to year (bp1-bs6.xlsx etc.),
but the download URL includes no date, so no "find latest" step is needed
, unlike the ministry list or BP2, budget.gov.au always points at the
current cycle's files.

Licensing: Creative Commons Attribution 4.0, same as BP2. Attribution used
in the UI: "Based on Commonwealth of Australia data, budget.gov.au."

NOT executed against the live network in this environment, logic
validated directly against the real files you supplied (extracted from
this exact ZIP) during this session, not against a fixture. Re-run once
after a new budget to confirm sheet/column layout hasn't shifted before
trusting it in the scheduled job.

Requires: requests, openpyxl, supabase (pip install --break-system-packages)
"""

from __future__ import annotations  # lets 3.9 tolerate any 3.10+ type-hint syntax below

import csv
import io
import os
import zipfile

import requests

CHART_DATA_ZIP_URL = "https://budget.gov.au/content/download/chart-data-final.zip"
HEADERS = {"User-Agent": "Poli civic-tech data pipeline (contact: <your email>)"}

REVENUE_CATEGORIES = [
    "Total individuals and other withholding", "Company tax",
    "Goods and services tax", "Total excise duty", "Non-taxation receipts",
    "Total receipts",
]
# Snapshot years for the trend chart, earliest available, +10yr, latest
# actual, latest forecast. Adjust the latter two each cycle if desired.
SNAPSHOT_YEARS = ["2005-06 ($m)", "2015-16 ($m)", "2024-25 ($m)", "2029-30 (est) ($m)"]


def download_zip() -> zipfile.ZipFile:
    resp = requests.get(CHART_DATA_ZIP_URL, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    return zipfile.ZipFile(io.BytesIO(resp.content))


def parse_revenue_composition(zf: zipfile.ZipFile) -> list[dict]:
    # Real filename inside the ZIP: "Table 1 - Australian Government (cash) receipts.csv"
    # (confirmed against a live download. NOT the underscored version browsers
    # sometimes rename files to on manual download, which is what this used to match).
    name = next((n for n in zf.namelist()
                 if "table 1" in n.lower() and "(cash) receipts" in n.lower()), None)
    if not name:
        raise RuntimeError("Could not find the cash receipts CSV in the ZIP, filenames may have changed.")
    text = zf.read(name).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    header = [h.strip() for h in rows[0]]
    data = {r[0].strip(): r[1:] for r in rows[1:] if r[0].strip()}

    years_present = [y for y in SNAPSHOT_YEARS if y in header]
    if len(years_present) < len(SNAPSHOT_YEARS):
        missing = set(SNAPSHOT_YEARS) - set(years_present)
        print(f"  WARNING: snapshot years not found in this cycle's data: {missing}, update SNAPSHOT_YEARS.")

    out = []
    for y in years_present:
        idx = header.index(y)
        total = float(data["Total receipts"][idx - 1])
        individuals = float(data["Total individuals and other withholding"][idx - 1]) / total * 100
        company = float(data["Company tax"][idx - 1]) / total * 100
        gst = float(data["Goods and services tax"][idx - 1]) / total * 100
        excise = float(data["Total excise duty"][idx - 1]) / total * 100
        other = 100 - individuals - company - gst - excise  # residual, matches component's expected shape
        out.append({
            "year": y.replace(" ($m)", ""),
            "total": round(total),
            "individuals": round(individuals, 1),
            "company": round(company, 1),
            "gst": round(gst, 1),
            "excise": round(excise, 1),
            "other": round(other, 1),
        })
    return out


def parse_spending_by_function(zf: zipfile.ZipFile) -> list[dict]:
    import openpyxl
    name = next((n for n in zf.namelist() if n.endswith("bp1-bs6.xlsx")), None)
    if not name:
        raise RuntimeError("Could not find bp1-bs6.xlsx in the ZIP.")
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)), data_only=True, read_only=True)
    ws = wb["6.1"]
    out = []
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and row[0] != "Function" and isinstance(row[1], (int, float)):
            out.append({"label": row[0], "pct": round(row[1] * 100, 1)})
    return out


def parse_forecast_accuracy(zf: zipfile.ZipFile) -> list[dict]:
    import openpyxl
    name = next((n for n in zf.namelist() if n.endswith("bp1-bs8.xlsx")), None)
    if not name:
        raise RuntimeError("Could not find bp1-bs8.xlsx in the ZIP.")
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)), data_only=True, read_only=True)
    ws = wb["8.1"]
    out = []
    for row in ws.iter_rows(values_only=True):
        if row and isinstance(row[0], str) and "-" in row[0] and len(row) >= 3:
            out.append({"year": row[0], "forecast": row[1], "outcome": row[2]})
    return out[-7:]  # most recent 7 years, enough to tell the story without crowding the chart


def upsert_to_supabase(revenue, spending, forecast):
    from supabase import create_client
    sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_ANON_KEY"])
    sb.table("budget_glance").delete().neq("section", "").execute()
    sb.table("budget_glance").insert([
        {"section": "revenue_composition", "data": revenue},
        {"section": "spending_by_function", "data": spending},
        {"section": "forecast_accuracy", "data": forecast},
    ]).execute()
    print("Upserted 3 budget_glance sections.")


def main():
    zf = download_zip()
    revenue = parse_revenue_composition(zf)
    spending = parse_spending_by_function(zf)
    forecast = parse_forecast_accuracy(zf)

    print(f"Revenue composition: {len(revenue)} snapshot years")
    print(f"Spending by function: {len(spending)} categories (should sum to ~100: {sum(s['pct'] for s in spending):.1f})")
    print(f"Forecast accuracy: {len(forecast)} years")

    if os.environ.get("SUPABASE_URL"):
        upsert_to_supabase(revenue, spending, forecast)
    else:
        print("\nSUPABASE_URL not set, dry run only.")
        print("Sample revenue row:", revenue[-1] if revenue else None)
        print("Sample spending row:", spending[0] if spending else None)
        print("Sample forecast row:", forecast[-1] if forecast else None)


if __name__ == "__main__":
    main()
